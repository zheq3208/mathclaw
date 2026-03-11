# -*- coding: utf-8 -*-
"""Recalculate Excel formulas via LibreOffice headless mode.

Uses a LibreOffice Basic macro to open, recalculate, and save. Then
validates the result with openpyxl to detect formula errors.

Usage:
    python recalc.py <excel_file> [timeout_seconds]
"""
from __future__ import annotations

import json
import os
import platform
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path


# ── LibreOffice macro ──────────────────────────────────────────────

_MACRO_CONTENT = """\
Sub RecalculateAndSave(sURL As String)
    Dim oDoc As Object
    Dim oSheets As Object
    Dim oSheet As Object
    Dim i As Long

    oDoc = StarDesktop.loadComponentFromURL( _
        ConvertToURL(sURL), "_blank", 0, Array())

    If IsNull(oDoc) Or IsEmpty(oDoc) Then
        MsgBox "Failed to open: " & sURL
        Exit Sub
    End If

    oSheets = oDoc.getSheets()
    For i = 0 To oSheets.getCount() - 1
        oSheet = oSheets.getByIndex(i)
        oSheet.calculateAll()
    Next i

    oDoc.store()
    oDoc.close(True)
End Sub
"""


def _install_macro() -> None:
    """Install the RecalculateAndSave macro into LibreOffice's user dir."""
    system = platform.system()
    if system == "Darwin":
        macro_dir = Path.home() / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    elif system == "Linux":
        macro_dir = Path.home() / ".config/libreoffice/4/user/basic/Standard"
    elif system == "Windows":
        macro_dir = Path.home() / "AppData/Roaming/LibreOffice/4/user/basic/Standard"
    else:
        raise RuntimeError(f"Unsupported platform: {system}")

    macro_dir.mkdir(parents=True, exist_ok=True)
    macro_file = macro_dir / "RecalculateAndSave.xba"
    macro_file.write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.'
        '0//EN" "module.dtd">\n'
        '<script:module xmlns:script="http://openoffice.org/2000/script" '
        'script:name="RecalculateAndSave" script:language="StarBasic">\n'
        f"{_MACRO_CONTENT}"
        "</script:module>\n"
    )


def _get_timeout_cmd() -> str:
    system = platform.system()
    if system == "Darwin":
        if shutil.which("gtimeout"):
            return "gtimeout"
        return "timeout"
    return "timeout"


# ── Excel error detection ─────────────────────────────────────────

_EXCEL_ERRORS = {
    "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#N/A", "#NUM!",
}


def _check_errors(filepath: str) -> dict:
    """Check for Excel formula errors using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed"}

    errors: list[dict] = []
    total_formulas = 0

    # data_only=True to read calculated values
    wb_data = openpyxl.load_workbook(filepath, data_only=True)
    # data_only=False to identify formula cells
    wb_formula = openpyxl.load_workbook(filepath, data_only=False)

    for sheet_name in wb_data.sheetnames:
        ws_d = wb_data[sheet_name]
        ws_f = wb_formula[sheet_name]
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1
                    data_cell = ws_d[cell.coordinate]
                    val = data_cell.value
                    if isinstance(val, str) and val.strip().upper() in _EXCEL_ERRORS:
                        errors.append(
                            {
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "formula": cell.value,
                                "error": val.strip(),
                            }
                        )

    wb_data.close()
    wb_formula.close()

    return {
        "total_formulas": total_formulas,
        "total_errors": len(errors),
        "errors": errors[:50],  # cap output
    }


# ── Main ───────────────────────────────────────────────────────────

def recalc(filename: str, timeout: int = 30) -> dict:
    """Recalculate formulas and return error report."""
    abs_path = os.path.abspath(filename)
    if not os.path.isfile(abs_path):
        return {"status": "error", "message": f"File not found: {abs_path}"}

    soffice = shutil.which("soffice")
    if not soffice:
        return {
            "status": "error",
            "message": "LibreOffice not found. Install with: brew install --cask libreoffice",
        }

    _install_macro()

    timeout_cmd = _get_timeout_cmd()
    cmd = [
        timeout_cmd,
        str(timeout),
        soffice,
        "--headless",
        "--invisible",
        f"macro:///Standard.RecalculateAndSave.RecalculateAndSave({abs_path})",
    ]

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=timeout + 10
        )
        if result.returncode != 0:
            return {
                "status": "warning",
                "message": f"soffice exit {result.returncode}: {result.stderr[:500]}",
            }
    except subprocess.TimeoutExpired:
        return {"status": "timeout", "message": f"Timed out after {timeout}s"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

    report = _check_errors(abs_path)
    report["status"] = "ok" if report["total_errors"] == 0 else "has_errors"
    return report


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    t = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    out = recalc(sys.argv[1], t)
    print(json.dumps(out, indent=2, ensure_ascii=False))
